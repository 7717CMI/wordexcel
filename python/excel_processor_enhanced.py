import os
import logging
import shutil
from typing import Dict, Any
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from config import Config

logger = logging.getLogger(__name__)

class ExcelProcessorEnhanced:
    def __init__(self, template_path: str):
        self.template_path = template_path
        self.workbook = None
        self.worksheet = None
        self.temp_path = None

    def load_template(self) -> None:
        """Load the Excel template with macro preservation"""
        try:
            logger.info(f'Attempting to load template from: {self.template_path}')
            
            if not os.path.exists(self.template_path):
                raise FileNotFoundError(f"Template file not found: {self.template_path}")
            
            logger.info(f'Template path exists: {os.path.exists(self.template_path)}')
            
            # Create a temporary copy to work with
            import tempfile
            temp_dir = tempfile.mkdtemp()
            self.temp_path = os.path.join(temp_dir, 'template_copy.xlsm')
            shutil.copy2(self.template_path, self.temp_path)
            
            logger.info(f'Template copied to temp location: {self.temp_path}')
            
            # Load workbook with VBA preservation
            self.workbook = load_workbook(self.temp_path, keep_vba=True)
            
            # Get the "Data" worksheet specifically
            if "Data" in self.workbook.sheetnames:
                self.worksheet = self.workbook["Data"]
            else:
                self.worksheet = self.workbook.active
            
            if not self.worksheet:
                raise Exception('No worksheet found in template')

            logger.info('Template loaded successfully')
            logger.info(f'Worksheet name: {self.worksheet.title}')
            logger.info(f'Total worksheets: {len(self.workbook.worksheets)}')
            
        except Exception as error:
            logger.error(f'Error loading template: {error}')
            raise Exception(f'Failed to load Excel template: {str(error)}')

    def populate_market_overview(self, data: Dict[str, Any]) -> None:
        """Populate market overview data on Data sheet (cells D2, D4-D10)"""
        try:
            logger.info('=== POPULATING MARKET OVERVIEW ===')
            
            # Validate data structure first
            if 'market' not in data:
                raise Exception("Missing 'market' key in data")
            
            market = data['market']
            logger.info(f'Market data keys: {list(market.keys())}')
            
            # Validate required market fields
            required_fields = ['market_name', 'size_base_raw', 'size_forecast_raw', 'cagr_percent_display', 'currency_unit']
            missing_fields = [field for field in required_fields if field not in market]
            if missing_fields:
                raise Exception(f"Missing required market fields: {missing_fields}")
            
            # Market Name - D2
            if not market.get('market_name'):
                raise Exception("Market name is empty or missing")
            self.worksheet['D2'] = market['market_name']
            logger.info(f'D2 (Market Name): {market["market_name"]}')

            # Base Year - D4 (SKIPPED - not populating)
            # self.worksheet['D4'] = market['base_year']
            logger.info(f'D4 (Base Year): SKIPPED - not populating')

            # Start Year - D5 (SKIPPED - not populating)
            # self.worksheet['D5'] = market['start_year']
            logger.info(f'D5 (Start Year): SKIPPED - not populating')

            # End Year - D6 (SKIPPED - not populating)
            # self.worksheet['D6'] = market['end_year']
            logger.info(f'D6 (End Year): SKIPPED - not populating')

            # Market Size Base Year - D7
            if not market.get('size_base_raw'):
                raise Exception("Market size base is empty or missing")
            self.worksheet['D7'] = market['size_base_raw']
            logger.info(f'D7 (Size Base): {market["size_base_raw"]}')

            # Market Size Forecast - D8
            if not market.get('size_forecast_raw'):
                raise Exception("Market size forecast is empty or missing")
            self.worksheet['D8'] = market['size_forecast_raw']
            logger.info(f'D8 (Size Forecast): {market["size_forecast_raw"]}')

            # CAGR - D9
            if not market.get('cagr_percent_display'):
                raise Exception("CAGR is empty or missing")
            cagr_value = market['cagr_percent_display']
            self.worksheet['D9'] = cagr_value
            logger.info(f'D9 (CAGR): {cagr_value}')

            # Currency Unit - D10
            if not market.get('currency_unit'):
                raise Exception("Currency unit is empty or missing")
            self.worksheet['D10'] = market['currency_unit']
            logger.info(f'D10 (Currency): {market["currency_unit"]}')

            # Driver 1 - C15
            if market.get('driver_1'):
                self.worksheet['C15'] = market['driver_1']
                logger.info(f'C15 (Driver 1): {market["driver_1"]}')
            else:
                logger.info('C15 (Driver 1): Not provided')

            # Driver 2 - C16
            if market.get('driver_2'):
                self.worksheet['C16'] = market['driver_2']
                logger.info(f'C16 (Driver 2): {market["driver_2"]}')
            else:
                logger.info('C16 (Driver 2): Not provided')

            # Restraint 1 - C17
            if market.get('restraint_1'):
                self.worksheet['C17'] = market['restraint_1']
                logger.info(f'C17 (Restraint 1): {market["restraint_1"]}')
            else:
                logger.info('C17 (Restraint 1): Not provided')

            # Restraint 2 - C18
            if market.get('restraint_2'):
                self.worksheet['C18'] = market['restraint_2']
                logger.info(f'C18 (Restraint 2): {market["restraint_2"]}')
            else:
                logger.info('C18 (Restraint 2): Not provided')

            logger.info('✅ Market overview populated successfully')
            
        except Exception as error:
            logger.error(f'Error populating market overview: {error}')
            logger.error(f'Data structure: {data}')
            raise Exception(f'Failed to populate market overview: {str(error)}')

    def populate_headers(self, data: Dict[str, Any]) -> None:
        """Populate dynamic headers (row 2) on Data sheet"""
        try:
            logger.info('=== POPULATING HEADERS ===')
            
            # Populate headers based on available segmentation data
            segments = data['segments']
            
            # Category 1 header - H2 (I column intentionally left empty)
            if segments.get('cat_1') and segments['cat_1'].get('header'):
                self.worksheet['H2'] = segments['cat_1']['header']
                logger.info(f'H2 Header: {segments["cat_1"]["header"]}')
            
            # Category 2 header - J2
            if segments.get('cat_2') and segments['cat_2'].get('header'):
                self.worksheet['J2'] = segments['cat_2']['header']
                logger.info(f'J2 Header: {segments["cat_2"]["header"]}')
            
            # Category 3 header - K2
            if segments.get('cat_3') and segments['cat_3'].get('header'):
                self.worksheet['K2'] = segments['cat_3']['header']
                logger.info(f'K2 Header: {segments["cat_3"]["header"]}')
            
            # Category 4 header - L2
            if segments.get('cat_4') and segments['cat_4'].get('header'):
                self.worksheet['L2'] = segments['cat_4']['header']
                logger.info(f'L2 Header: {segments["cat_4"]["header"]}')
            
            # Category 5 header - M2
            if segments.get('cat_5') and segments['cat_5'].get('header'):
                self.worksheet['M2'] = segments['cat_5']['header']
                logger.info(f'M2 Header: {segments["cat_5"]["header"]}')
            
            logger.info('✅ Headers populated successfully')
            
        except Exception as error:
            logger.error(f'Error populating headers: {error}')
            raise Exception('Failed to populate headers')

    def populate_lists(self, data: Dict[str, Any]) -> None:
        """Populate dynamic lists (rows 3+) on Data sheet"""
        try:
            logger.info('=== POPULATING LISTS ===')
            
            # Clear existing data in rows 3-80 (handle merged cells). Columns: H, J, K, L, M. Column I left empty.
            for row in range(3, 81):
                for col in ['H', 'J', 'K', 'L', 'M']:  # Columns H, J, K, L, M
                    cell = self.worksheet[f'{col}{row}']
                    # Check if cell is part of a merged range
                    if hasattr(cell, 'value') and not isinstance(cell, type(None)):
                        try:
                            cell.value = None
                        except AttributeError:
                            # Skip merged cells - they're read-only
                            logger.debug(f'Skipping merged cell at {col}{row}')
                            continue
            
            # Segmentation category 1 items - Column H, starting from row 3
            if data['segments'].get('cat_1') and data['segments']['cat_1'].get('items'):
                for index, item in enumerate(data['segments']['cat_1']['items']):
                    if index < 71:
                        row = 3 + index
                        self.worksheet[f'H{row}'] = f">{item}"

            # Segmentation category 2 items - Column J, starting from row 3
            if data['segments'].get('cat_2') and data['segments']['cat_2'].get('items'):
                for index, item in enumerate(data['segments']['cat_2']['items']):
                    if index < 71:
                        row = 3 + index
                        self.worksheet[f'J{row}'] = f">{item}"

            # Segmentation category 3 items - Column K, starting from row 3
            if data['segments'].get('cat_3') and data['segments']['cat_3'].get('items'):
                for index, item in enumerate(data['segments']['cat_3']['items']):
                    if index < 71:
                        row = 3 + index
                        self.worksheet[f'K{row}'] = f">{item}"

            # Segmentation category 4 items - Column L, starting from row 3
            if data['segments'].get('cat_4') and data['segments']['cat_4'].get('items'):
                for index, item in enumerate(data['segments']['cat_4']['items']):
                    if index < 71:
                        row = 3 + index
                        self.worksheet[f'L{row}'] = f">{item}"

            # Segmentation category 5 items - Column M, starting from row 3
            if data['segments'].get('cat_5') and data['segments']['cat_5'].get('items'):
                for index, item in enumerate(data['segments']['cat_5']['items']):
                    if index < 71:
                        row = 3 + index
                        self.worksheet[f'M{row}'] = f">{item}"

            logger.info('✅ Lists populated successfully')
            
        except Exception as error:
            logger.error(f'Error populating lists: {error}')
            raise Exception('Failed to populate lists')

    def populate_key_players(self, data: Dict[str, Any]) -> None:
        """Populate key players data on Data sheet (Column G, rows 3-30)"""
        try:
            logger.info('=== POPULATING KEY PLAYERS ===')
            
            players = data.get('players', {})
            player_list = players.get('players', [])
            
            # Clear existing data in column G, rows 3-30
            for row in range(3, 31):
                cell = self.worksheet[f'G{row}']
                if hasattr(cell, 'value') and not isinstance(cell, type(None)):
                    try:
                        cell.value = None
                    except AttributeError:
                        # Skip merged cells - they're read-only
                        logger.debug(f'Skipping merged cell at G{row}')
                        continue
            
            # Populate key players starting from G3
            for index, player in enumerate(player_list):
                if index < 28:  # H3 to H30 (28 players max)
                    row = 3 + index
                    self.worksheet[f'G{row}'] = player
                    logger.info(f'G{row} Player: {player}')
            
            logger.info(f'✅ Key players populated successfully: {len(player_list)} players')
            
        except Exception as error:
            logger.error(f'Error populating key players: {error}')
            raise Exception('Failed to populate key players')

    def populate_data(self, data: Dict[str, Any]) -> None:
        """Main method to populate all data"""
        try:
            self.load_template()
            self.populate_market_overview(data)
            self.populate_headers(data)
            self.populate_lists(data)
            self.populate_key_players(data)
            logger.info('All data populated successfully')
        except Exception as error:
            logger.error(f'Error populating data: {error}')
            raise error

    def save_file(self, output_path: str) -> None:
        """Save the populated Excel file with macro preservation using openpyxl"""
        try:
            logger.info(f'Saving Excel file to: {output_path}')
            
            # Ensure output directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Save using openpyxl with macro preservation
            self.workbook.save(output_path)
            logger.info('✅ File saved using openpyxl (macro preservation)')
            
            # Verify the file
            self._verify_file(output_path)
                
        except Exception as error:
            logger.error(f'Error saving Excel file: {error}')
            raise Exception(f'Failed to save Excel file: {str(error)}')

    def _verify_file(self, output_path: str) -> None:
        """Verify the saved file is valid"""
        if os.path.exists(output_path):
            stats = os.stat(output_path)
            logger.info(f'Excel file saved successfully to: {output_path}')
            logger.info(f'File size: {stats.st_size} bytes')
            
            # Verify it's a valid Excel file
            with open(output_path, 'rb') as f:
                file_signature = f.read(4).hex()
            
            if not file_signature.startswith('504b'):
                logger.error('ERROR: File does not appear to be a valid Excel file')
                raise Exception('Generated file is not a valid Excel file')
            else:
                logger.info('✅ File signature validation passed')
        else:
            raise Exception('File was not created')

    def cleanup(self):
        """Clean up temporary files"""
        try:
            if self.temp_path and os.path.exists(self.temp_path):
                os.remove(self.temp_path)
                logger.info('Temporary file cleaned up')
        except Exception as e:
            logger.warning(f'Cleanup warning: {e}')

    def __del__(self):
        """Destructor to ensure cleanup"""
        self.cleanup()
